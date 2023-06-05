import cv2
import dlib
import time
import threading
import math
import datetime
import os
import openpyxl
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import numpy as np
from skimage import util
from skimage.io import imread
from skimage.filters import threshold_otsu
from skimage import measure
from skimage.measure import regionprops
import matplotlib.patches as patches
import sklearn
import joblib

SNo = int(1)
counter = 0
sheet = {}
now = datetime.datetime.now()

# mainFilename = 'database.xlsx'

# if not(os.path.isfile(mainFilename)):
#     mwb = openpyxl.Workbook() #main work book
#     #main_sheet = mwb.create_sheet('main_sheet')
#     #main_sheet = mwb.worksheets
#     main_sheet=mwb.active  
#     #main_sheet = main_sheet.title
#     main_sheet.cell(row=1, column=1).value = 'VehicleID'
#     main_sheet.cell(row=1, column=2).value = 'Vehicle_Number'


#     sheet[0] = mwb.create_sheet('sheet%d' %counter)
#     # ws[counter] = mwb.worksheets[counter]
#     sheet[0] = mwb.active
#     a = 'Daily Vehicles'
#     day = now.day
#     b = str(str(a) + str(day) + '.' + str(now.month) + '.' + str(now.year))
#     sheet[0].title = b
#     sheet[0].cell(row=1, column=1).value = 'VehicleID'
#     sheet[0].cell(row=1, column=2).value = 'Date'
#     sheet[0].cell(row=1, column=3).value = 'Time'
#     sheet[0].cell(row=1, column=4).value = 'Speed'

#     mwb.save(mainFilename)

carCascade = cv2.CascadeClassifier('myhaar.xml')
video = cv2.VideoCapture('car16.mp4')

WIDTH = 1280
HEIGHT = 720

def detection(img_1):
    height = 900
    width = 700
    img = cv2.cvtColor(img_1, cv2.COLOR_RGB2GRAY)
    car_image = cv2.resize(img, (width, height))
    img_1 = cv2.resize(img, (width, height))
    
    #in skimage ranges between 0 & 1
    #multiplying it with 255 will make it range between 0 & 255 which we can relate better 

    gray_car_image = car_image * 255
    #fig, (ax1, ax2) = plt.subplots(1, 2)
    #ax1.imshow(gray_car_image, cmap="gray")
    # ax1.imshow(car_image, cmap="gray")
    threshold_value = threshold_otsu(gray_car_image)
    binary_car_image = gray_car_image > threshold_value
    # ax2.imshow(binary_car_image, cmap="gray")
    # plt.show()
    
    # Applying CCA to find connected regions of binary image
    #this gets all the connected regions and groups them together
    label_image = measure.label(binary_car_image)

    # print(label_image.shape[0]) height of the image
    # print(label_image.shape[1]) width of the image

    #getting the maximum width, height and minimum width and height that a license plate can be
    plate_dimensions = (0.03*label_image.shape[0], 0.08*label_image.shape[0], 0.15*label_image.shape[1], 0.3*label_image.shape[1])
    plate_dimensions2 = (0.08*label_image.shape[0], 0.2*label_image.shape[0], 0.15*label_image.shape[1], 0.4*label_image.shape[1])
    #plate_dimensions = (0.06*label_image.shape[0], 0.13*label_image.shape[0], 0.085*label_image.shape[1], 0.2*label_image.shape[1])
    #plate_dimensions2 = (0.06*label_image.shape[0], 0.18*label_image.shape[0], 0.15*label_image.shape[1], 0.25*label_image.shape[1])

    # plate_dimensions = (0.03*label_image.shape[0], 0.08*label_image.shape[0], 0.15*label_image.shape[1], 0.3*label_image.shape[1])
    # plate_dimensions2 = (0.08*label_image.shape[0], 0.2*label_image.shape[0], 0.15*label_image.shape[1], 0.4*label_image.shape[1])
    min_height, max_height, min_width, max_width = plate_dimensions
    plate_objects_cordinates = []
    plate_like_objects = []

    #print(min_height, ' ', max_height, ' ', min_width, ' ', max_width)

    # fig, (ax1) = plt.subplots(1)
    # ax1.imshow(gray_car_image, cmap="gray")

    flag = 0
    go = 0

    #regionprops creates a list of properties of all the labelled regions
    for region in regionprops(label_image):

        #min_row, min_col, max_row, max_col = region.bbox
        #print('')
        #print('min_row:', min_row, 'max_row:', max_row, 'min_col:', min_col, 'max_col:', max_col)
        #print(region.area)
        #print('')

        if (region.area < 50):
            #if the region is so small then it's likely not a license plate
            continue
        
        #the bounding box coordinates
        # print(type(region.bbox))
        # print(region.bbox)

        min_row, min_col, max_row, max_col = region.bbox
        if min_col<100:
             continue
        region_height = max_row - min_row
        region_width = max_col - min_col

        #ensuring that the region identified satisfies the condition of a typical license plate
        if(region_height >= min_height and region_height <= max_height and region_width >= min_width and region_width <= max_width and region_width > region_height):

            flag = 1
            go = go + 1
            # print(go)
            fig, (ax1) = plt.subplots(1)
            ax1.imshow(gray_car_image, cmap="gray")
        
            plate_like_objects.append(binary_car_image[min_row:max_row, min_col:max_col])
            plate_objects_cordinates.append((min_row, min_col, max_row, max_col))
            rectBorder = patches.Rectangle((min_col, min_row), max_col - min_col, max_row - min_row, edgecolor = "red", linewidth = 2, fill = False )
            ax1.add_patch(rectBorder)
            plt.show()
            
            #cropped_image = gray_car_image[min_row:max_row, min_col:max_col]
            #print(type(cropped_image))
            #changed_color = cv2.cvtColor(cropped_image, cv2.COLOR_BGR2GRAY)
    
            #inverted_image = util.invert(cropped_image)
            #threshold_value_license_plate = threshold_otsu(inverted_image)
            #print(threshold_value_license_plate)
            #thresholded_image = inverted_image > threshold_value_license_plate
            
            #fig, (ax1) = plt.subplots(1)
            #ax1.imshow(thresholded_image, cmap='gray')
            #plt.show()
            
            # let's draw a red rectangle over those regions

    # if (flag == 1):
    #     print(plate_like_objects[0])
        #plt.show()
        #print('flag=1')
    
    if (flag == 0):
        min_height, max_height, min_width, max_width = plate_dimensions2
        plate_objects_cordinates = []
        plate_like_objects = []

        #fig, (ax1) = plt.subplots(1)
        #ax1.imshow(gray_car_image, cmap="gray")

        # regionprops creates a list of properties of all the labelled regions
        for region in regionprops(label_image):
            if (region.area < 150):
                #if the region is so small then it's likely not a license plate  
                continue
            #bounding box coordinates
            min_row, min_col, max_row, max_col = region.bbox
            if min_col<100:
               continue
            region_height = max_row - min_row
            region_width = max_col - min_col

            # ensuring that the region identified satisfies the condition of a typical license plate
            if(region_height >= min_height and region_height <= max_height and region_width >= min_width and region_width <= max_width and region_width > region_height):
                fig, (ax1) = plt.subplots(1)
                ax1.imshow(gray_car_image, cmap="gray")
                
                plate_like_objects.append(binary_car_image[min_row:max_row, min_col:max_col])
                plate_objects_cordinates.append((min_row, min_col, max_row, max_col))
                rectBorder = patches.Rectangle((min_col, min_row), max_col - min_col, max_row - min_row, edgecolor="red", linewidth=2, fill=False)
                ax1.add_patch(rectBorder)
                plt.show()
                #print('flag=0')
                
                #cropped_image = gray_car_image[min_row:max_row, min_col:max_col]
                #print(type(cropped_image))
                #changed_color = cv2.cvtColor(cropped_image, cv2.COLOR_BGR2GRAY)
        
                #inverted_image = util.invert(cropped_image)
                #fig, (ax1) = plt.subplots(1)
                #ax1.imshow(inverted_image, cmap='gray')
                #plt.show()
                # let's draw a red rectangle over those regions

        #print(plate_like_objects[3])
        #plt.show()   
    
    #print(flag)
    #print(len(plate_like_objects))
    #print(plate_like_objects[len(plate_like_objects)-1])
    #print(plate_objects_cordinates[len(plate_like_objects)-1])

    if(len(plate_like_objects) != 0):
        print(len(plate_like_objects))
        (r1, c1, r2, c2) = plate_objects_cordinates[len(plate_like_objects)-1]
        #cropped_image_1 = gray_car_image[r1:r2, c1:c2]
        img_2 = img_1[r1:r2, c1:c2]
        img_3 = cv2.medianBlur(img_2, 5)

        ret, th_1 = cv2.threshold(img_3, 155, 255, cv2.THRESH_BINARY)
        th_2 = cv2.adaptiveThreshold(img_3, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 11, 2)
        not_th_2 = cv2.bitwise_not(th_2)

        width_plate = 180
        # 90
        height_plate = 150
        # 60

        license_plate = cv2.resize(not_th_2, (width_plate, height_plate)) 
        #--------------Display license plate ------------------------------------
        # cv2.imshow('license plate', license_plate)
        # cv2.waitKey(0)
        # fig, (ax1) = plt.subplots(1)
        # ax1.imshow(license_plate, cmap="gray")


# def estimateSpeed(location1, location2):
def estimateSpeed(location1, location2):
    d_pixels = math.sqrt(math.pow(location2[0]-location1[0],2)+ math.pow(location2[1]-location1[1],2))
    ppm = 10
    d_meters = d_pixels/ppm
    fps = 7
    # 10
    speed = d_meters * fps * 3.6 #3.6 to convert m/s to km/hr
    return speed

def trackMultipleObjects():
    rectangleColor = (0, 255, 0)
    frameCounter = 0
    currentCarID = 0
    fps = 0
    
    cropped_image ={}
    carTracker = {}
    carNumbers = {}
    carLocation1 = {}
    carLocation2 = {}
    speed = [None]*1000

    # Write output to video file
    out = cv2.VideoWriter('output.avi', cv2.VideoWriter_fourcc('M', 'J', 'P', 'G'), 10, (WIDTH, HEIGHT))

    while True:
        start_time = time.time()
        rc, image_2 = video.read()

        if type(image_2) == type(None):
            break
        image_1 = cv2.resize(image_2, (WIDTH, HEIGHT))
        resultImage = image_1.copy()
        height = image_1.shape[0]
        width = image_1.shape[1]
        image = image_1[0:height, 0:int(width/2)]

        frameCounter = frameCounter + 1
        carIDtoDelete = []

        for carID in carTracker.keys():
            trackingQuality = carTracker[carID].update(image)

            if trackingQuality < 7:
                carIDtoDelete.append(carID)

        for carID in carIDtoDelete:
            print("Removing" + str(carID)+ "from list of trackers")
            print("Removing"+ str(carID)+ "previous location")
            print("Removing"+ str(carID)+ "current location")
            carTracker.pop(carID, None)
            carLocation1.pop(carID, None)
            carLocation2.pop(carID, None)
        
        if not (frameCounter % 10):
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
            cars = carCascade.detectMultiScale(gray, 1.1, 13, 18, (24, 24))

            for (_x, _y, _w, _h) in cars:
                x = int(_x)
                y = int(_y)
                w = int(_w)
                h = int(_h)

                x_bar = x + 0.5 * w
                y_bar = y + 0.5 * h

                matchCarID = None

                for carID in carTracker.keys():
                    trackedPosition = carTracker[carID].get_position()

                    t_x = int(trackedPosition.left())
                    t_y = int(trackedPosition.top())
                    t_w = int(trackedPosition.width())
                    t_h = int(trackedPosition.height())

                    t_x_bar = t_x + 0.5 * t_w
                    t_y_bar = t_y + 0.5 * t_h

                    if ((t_x <= x_bar <= (t_x + t_w)) and (t_y <= y_bar <= (t_y + t_h)) and (x <= t_x_bar <= (x + w)) and (y <= t_y_bar <= (y + h))):
                        
                        matchCarID = carID

                if matchCarID is None:
                    print("Creating new tracker: "+ str(currentCarID))

                    tracker = dlib.correlation_tracker()
                    tracker.start_track(image, dlib.rectangle(x, y, x + w, y + h))

                    carTracker[currentCarID] = tracker
                    carLocation1[currentCarID] = [x, y, w, h]

                    currentCarID = currentCarID + 1
                    
        for carID in carTracker.keys():
            trackedPosition = carTracker[carID].get_position() 

            t_x = int(trackedPosition.left())
            t_y = int(trackedPosition.top())
            t_w = int(trackedPosition.width())
            t_h = int(trackedPosition.height())

            cv2.rectangle(resultImage, (t_x, t_y), (t_x + t_w, t_y + t_h), rectangleColor, 4)

            carLocation2[carID] = [t_x, t_y, t_w, t_h] 

        end_time = time.time()

        if not (end_time == start_time):
            fps = 1.0/(end_time - start_time)
        length=len(carLocation1.keys())
        for a,i in enumerate(carLocation1.keys()):
            if a<length-1:
              continue
            if frameCounter % 1 == 0:
                [x1, y1, w1, h1] = carLocation1[i]
                [x2, y2, w2, h2] = carLocation2[i]

                carLocation1[i] = [x2, y2, w2, h2]

                if [x1, y1, w1, h1] != [x2, y2, w2, h2]:
                    if (speed[i] == None or speed[i] == 0):
                        speed[i] = estimateSpeed([x1, y1, w1, h1], [x2, y2, w2, h2])
                    if x1<0 and x2<0:
                        continue
                    if speed[i] != None and y1 >= 180:
                        cv2.putText(resultImage, str(int(speed[i])) + "km/hr", (int(x1 + w1/2), int(y1-5)), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (255, 255, 255), 2)

                    # if speed[i] > 6:
                    #     cropped_image[i] = image[y1:y1+h1, x1:x1+w1]
                    #     # fig, (ax1) = plt.subplots(1)
                    #     # ax1.imshow(cropped_image[i])
                    #     # plt.show()
                                        
                    #     # cv2.imshow('asghsa',cropped_image[i])
                    #     detection(cropped_image[i])
                        #print(type(cropped_image[i]))
                        #print(cropped_image[i])
                    
        cv2.imshow('result', resultImage)

        

        #write the frame into the file output.avi
        # out.write(resultImage)

        if cv2.waitKey(1) == 27:
            break
    
    cv2.destroyAllWindows()
    # out.release()

if __name__ == '__main__':

    trackMultipleObjects()





         

        







